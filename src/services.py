import os.path
import random
import string

import aiofiles as aiofiles
from fastapi import File
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel

from src.config import TEMP_DIR


class ExcelColumn(BaseModel):
    sheet_name: str
    column_name: str


def get_column_location(name: str, workbook: load_workbook) -> ExcelColumn:
    for sheet in workbook.sheetnames:
        for row in workbook[sheet].rows:
            for cell in row:
                if cell.value == name:
                    return ExcelColumn(sheet_name=sheet, column_name=cell.column_letter)


def get_x_value(
    workbook: Workbook, before_column_info: ExcelColumn, after_column_info: ExcelColumn
) -> str:
    values_dict = {}
    counter = 2
    exit_mark = False
    while exit_mark == False:
        before_cell = workbook[before_column_info.sheet_name][
            before_column_info.column_name + str(counter)
        ].value
        after_cell = workbook[after_column_info.sheet_name][
            after_column_info.column_name + str(counter)
        ].value
        if before_cell != None:
            try:
                values_dict[before_cell] += 1
            except KeyError:
                values_dict[before_cell] = 1
        else:
            exit_mark = True
            result = "added: "
        if after_cell != None:
            try:
                values_dict[after_cell] += 1
            except KeyError:
                values_dict[before_cell] = 1
        else:
            exit_mark = True
            result = "removed: "
        counter += 1

    x_value = list(values_dict.keys())[list(values_dict.values()).index(1)]
    return result + str(x_value)


def parser_excel(name_file: str) -> str:
    workbook = load_workbook(name_file)

    before_column = get_column_location("before", workbook)
    after_column = get_column_location("after", workbook)
    x_value = get_x_value(workbook, before_column, after_column)
    return x_value


async def save_file(file: File) -> str:
    random_name = "".join([random.choice(string.ascii_letters) for _ in range(5)])
    filepath = os.path.join(TEMP_DIR, f"{random_name}.xlsx")

    async with aiofiles.open(filepath, "wb") as out_file:
        while content := await file.read(1024):
            await out_file.write(content)

    return filepath
